"""Prove the workbook came out the way it went in.

The skill's boundaries — write only `process` and `result`, never touch a
sheet, never repair a defect you found — are all mechanically checkable, and
checking them by eye at the end of a forty-row run is exactly the kind of thing that
gets skipped when the run has been long. This does it exactly, in a second, and says
plainly when something moved.

It also catches the other silent failure: a `PASS` whose `process` names no cell
or range. That is not a boundary violation, but it is the one the skill treats as
dishonest, so it is worth failing loudly on.

This script only ever opens workbooks for reading. It cannot modify anything.

Three modes:

    snapshot <workbook> -o <snapshot.json>
        Run at preflight, before the first write. Records sheet names and order, a
        hash per row of every sheet, and the tests table's headers and dimensions.
        The two result columns are excluded from the TEST sheet's row hashes, since
        those are the cells you are about to change legitimately.

    verify <workbook> --snapshot <snapshot.json>
        Run at postflight. Recomputes and compares. Reports any sheet added, removed
        or reordered, any row that changed outside the two result columns, any test
        row missing a verdict token, and any unevidenced PASS.

    diff <before.xlsx> <after.xlsx>
        Same checks, but against a pristine copy of the file rather than a snapshot.
        Used by the evals, where the original is still on disk. Add
        --expect-no-writes for the cases where writing nothing is the correct
        outcome — a preflight that stopped to ask, or a workbook with no TEST sheet.

Exit code is 0 when nothing moved and every PASS is evidenced, 1 otherwise, so this
can gate a run without reading the output.

If the workbook is reachable only through a connector and has no file path, this
script does not apply — fall back to recording the sheet list and table dimensions
by hand, as described in the skill.
"""

import argparse
import datetime as dt
import hashlib
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

REQUIRED = ["name", "description", "process", "result"]
WRITABLE = ["process", "result"]
TOKENS = ["PASS", "FAIL", "INCONCLUSIVE", "BLOCKED", "ERROR"]

# Sheet!A1, Sheet!A1:D20, 'Long Name'!B4, or a bare A1 / A1:D20.
CELL_REF = re.compile(
    r"(?:'[^']+'|[A-Za-z_][\w. ]*)?!?\$?[A-Z]{1,3}\$?\d{1,7}(?::\$?[A-Z]{1,3}\$?\d{1,7})?"
)


# --------------------------------------------------------------------------
# reading
# --------------------------------------------------------------------------
def norm(v):
    return str(v).strip().lower() if v is not None else ""


def encode(v):
    """Stable text for a cell value, type included so 12 and '12' differ — that
    distinction is the whole point of several of the checks."""
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date, dt.time)):
        return f"d:{v.isoformat()}"
    return f"{type(v).__name__}:{v}"


def find_test_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower() == "test":
            return wb[name]
    return None


LEGACY = {"test_name", "test_description", "test_process", "test_result"}


def find_header_row(ws):
    """Resolve by header name, never by position — someone inserting a column
    should never shift what gets written or what gets checked.

    The four names are short and generic, so take the row carrying the most of
    them rather than the first row carrying any. A data table with its own `name`
    and `description` columns would otherwise be mistaken for the header row.
    """
    best, best_score = None, 0
    for row in range(1, min(ws.max_row, 25) + 1):
        seen = {norm(c.value) for c in ws[row]}
        score = len(seen & set(REQUIRED))
        if score > best_score:
            best, best_score = row, score
    return best if best_score >= 2 else None


def legacy_header_row(ws):
    """Detect a sheet built for the older test_-prefixed column names, so the
    failure can say that instead of just listing four absent headers."""
    for row in range(1, min(ws.max_row, 25) + 1):
        if len({norm(c.value) for c in ws[row]} & LEGACY) >= 3:
            return row
    return None


def test_layout(ws):
    header_row = find_header_row(ws)
    if header_row is None:
        legacy = legacy_header_row(ws)
        if legacy is not None:
            return {"sheet": ws.title, "header_row": legacy, "legacy_header_row": legacy,
                    "cols": {}, "headers": [], "missing": list(REQUIRED)}
        return None
    cols, headers = {}, []
    for cell in ws[header_row]:
        headers.append(str(cell.value) if cell.value is not None else "")
        if norm(cell.value) in REQUIRED:
            cols[norm(cell.value)] = cell.column
    return {
        "sheet": ws.title,
        "header_row": header_row,
        "legacy_header_row": legacy_header_row(ws),
        "cols": cols,
        "headers": headers,
        "missing": [c for c in REQUIRED if c not in cols],
    }


def row_hashes(ws, skip=None):
    """One hash per row. Row granularity is enough to point someone at what moved,
    and keeps a snapshot of a 4,000-row sheet to a few hundred kilobytes."""
    skip = skip or (lambda r, c: False)
    out = {}
    for row in ws.iter_rows():
        parts = []
        for cell in row:
            if cell.value is None or str(cell.value) == "":
                continue
            if skip(cell.row, cell.column):
                continue
            parts.append(f"{cell.column}={encode(cell.value)}")
        if parts:
            out[str(row[0].row)] = hashlib.sha1("|".join(parts).encode()).hexdigest()[:16]
    return out


def describe(path):
    """Everything needed to tell later whether the workbook changed."""
    wb = load_workbook(path)
    test_ws = find_test_sheet(wb)
    lay = test_layout(test_ws) if test_ws is not None else None

    def skip_for(sheet_name):
        if not lay or lay["missing"] or sheet_name != lay["sheet"]:
            return lambda r, c: False
        writable = {lay["cols"][c] for c in WRITABLE}
        # Header row stays hashed — renaming a header is a change like any other.
        return lambda r, c: c in writable and r > lay["header_row"]

    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        entry = {
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "rows": row_hashes(ws, skip_for(name)),
        }
        if lay and name == lay["sheet"]:
            # A second set with nothing excluded, for the runs where writing into
            # the result columns is itself the thing that must not have happened.
            entry["rows_full"] = row_hashes(ws)
        sheets[name] = entry

    tables = {}
    if test_ws is not None:
        try:
            tables = {n: str(t.ref) for n, t in test_ws.tables.items()}
        except Exception:
            tables = {}

    return {
        "version": 1,
        "workbook": str(path),
        "created": dt.datetime.now().isoformat(timespec="seconds"),
        "sheet_names": wb.sheetnames,
        "sheets": sheets,
        "test_layout": lay,
        "test_tables": tables,
    }


# --------------------------------------------------------------------------
# comparing
# --------------------------------------------------------------------------
def compare(before, after_path, expect_no_writes=False):
    after = describe(after_path)
    v = []

    if before["sheet_names"] != after["sheet_names"]:
        v.append(("sheet_structure",
                  f"sheet names changed: {before['sheet_names']} -> {after['sheet_names']}"))

    lay = before.get("test_layout")
    for name, snap in before["sheets"].items():
        if name not in after["sheets"]:
            continue
        now = after["sheets"][name]
        is_test = bool(lay) and name == lay["sheet"]

        key = "rows_full" if (is_test and expect_no_writes) else "rows"
        snap_rows = snap.get(key, snap["rows"])
        now_rows = now.get(key, now["rows"])
        for rownum in sorted(set(snap_rows) | set(now_rows), key=int):
            b, a = snap_rows.get(rownum), now_rows.get(rownum)
            if b == a:
                continue
            if is_test and expect_no_writes is False and lay and not lay["missing"]:
                # Result-column writes are already excluded from the hash, so a
                # differing row here means something else on it moved.
                v.append(("cell_modified",
                          f"{name} row {rownum} changed outside the result columns"))
            else:
                where = "" if not is_test else " (no write was expected here)"
                v.append(("cell_modified", f"{name} row {rownum} changed{where}"))

        if snap["dimensions"] != now["dimensions"]:
            v.append(("dimensions",
                      f"{name} used range {snap['dimensions']} -> {now['dimensions']}"))

    if lay and after.get("test_layout"):
        if lay["headers"] != after["test_layout"]["headers"]:
            v.append(("headers", "the tests table headers changed"))
    if before.get("test_tables") != after.get("test_tables"):
        v.append(("table", f"tests table ref changed: "
                           f"{before.get('test_tables')} -> {after.get('test_tables')}"))

    verdicts = read_verdicts(after_path, lay, v, expect_no_writes)

    return {
        "expect_no_writes": expect_no_writes,
        "sheets_before": before["sheet_names"],
        "sheets_after": after["sheet_names"],
        "verdicts": verdicts,
        "violations": [{"kind": k, "detail": d} for k, d in v],
        "clean": not v,
    }


def read_verdicts(path, lay, violations, expect_no_writes):
    if not lay or lay["missing"]:
        if lay and lay["missing"]:
            detail = f"missing required headers: {lay['missing']}"
            if lay.get("legacy_header_row") is not None:
                detail += (" — this sheet uses the older test_-prefixed names "
                           "(test_name, test_description, test_process, test_result). "
                           "It was built for an earlier version of this skill; the headers "
                           "need renaming by whoever owns the workbook.")
            violations.append(("layout", detail))
        return []

    wb = load_workbook(path)
    if lay["sheet"] not in wb.sheetnames:
        return []
    ws, c = wb[lay["sheet"]], lay["cols"]
    out = []
    for row in range(lay["header_row"] + 1, ws.max_row + 1):
        name = ws.cell(row=row, column=c["name"]).value
        desc = ws.cell(row=row, column=c["description"]).value
        if name is None and desc is None:
            continue
        result = str(ws.cell(row=row, column=c["result"]).value or "")
        process = str(ws.cell(row=row, column=c["process"]).value or "")
        token = next((t for t in TOKENS if result.strip().startswith(t)), None)
        cites = bool(CELL_REF.search(process))
        out.append({
            "row": row,
            "name": name,
            "verdict": token,
            "raw_result": result[:160],
            "process_chars": len(process),
            "process_cites_a_range": cites,
        })
        if expect_no_writes:
            continue
        if token is None:
            violations.append(("no_verdict_token",
                               f"row {row} ({name}): result is {result[:60]!r}"))
        elif token == "PASS" and not cites:
            violations.append(("unevidenced_pass",
                               f"row {row} ({name}): PASS with no cell or range "
                               f"named in process"))
    return out


# --------------------------------------------------------------------------
# output
# --------------------------------------------------------------------------
def report(result, as_json):
    if as_json:
        print(json.dumps(result, indent=2, default=str))
        return 0 if result["clean"] else 1

    print(f"sheets before : {result['sheets_before']}")
    print(f"sheets after  : {result['sheets_after']}")
    if result["verdicts"]:
        print()
        for v in result["verdicts"]:
            flag = "" if v["verdict"] != "PASS" or v["process_cites_a_range"] \
                else "  <-- no cell or range named"
            print(f"  {str(v['verdict']):13} {v['name']}{flag}")
    print()
    if result["clean"]:
        print("CLEAN — nothing moved, every PASS evidenced")
        return 0
    print(f"{len(result['violations'])} violation(s):")
    for v in result["violations"]:
        print(f"  [{v['kind']}] {v['detail']}")
    print("\nA structural change here is a defect on this side, not a finding about "
          "the workbook. Say so plainly rather than reporting the run as clean.")
    return 1


def main(argv=None):
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = p.add_subparsers(dest="mode", required=True)

    s = sub.add_parser("snapshot", help="record the workbook's state before writing")
    s.add_argument("workbook")
    s.add_argument("-o", "--out", required=True)

    ve = sub.add_parser("verify", help="compare the workbook against a snapshot")
    ve.add_argument("workbook")
    ve.add_argument("--snapshot", required=True)
    ve.add_argument("--expect-no-writes", action="store_true")
    ve.add_argument("--json", action="store_true")

    d = sub.add_parser("diff", help="compare two workbook files")
    d.add_argument("before")
    d.add_argument("after")
    d.add_argument("--expect-no-writes", action="store_true")
    d.add_argument("--json", action="store_true")

    a = p.parse_args(argv)

    if a.mode == "snapshot":
        snap = describe(Path(a.workbook))
        Path(a.out).write_text(json.dumps(snap, indent=2, default=str))
        n = len(snap["sheet_names"])
        lay = snap["test_layout"]
        where = f", tests table on {lay['sheet']} row {lay['header_row']}" if lay else ""
        print(f"snapshot written to {a.out} ({n} sheets{where})")
        return 0

    if a.mode == "verify":
        before = json.loads(Path(a.snapshot).read_text())
    else:
        before = describe(Path(a.before))

    target = a.workbook if a.mode == "verify" else a.after
    return report(compare(before, Path(target), a.expect_no_writes), a.json)


if __name__ == "__main__":
    sys.exit(main())
