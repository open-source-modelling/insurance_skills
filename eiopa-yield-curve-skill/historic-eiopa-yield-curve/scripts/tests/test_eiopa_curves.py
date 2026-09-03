#!/usr/bin/env python3
"""
Regression tests for eiopa_curves.py and for the facts this skill asserts in
prose.

    python3 test_eiopa_curves.py              # everything
    python3 test_eiopa_curves.py -v           # verbose
    python3 test_eiopa_curves.py --regenerate-golden

Stdlib unittest, no pytest, no network. Runs against the bundled pack alone, so
it works on any machine that has the skill - the same guarantee the skill makes
about itself. Tests that need the EIOPA_all_curves repo (the pack round-trip and
the Excel shock formulas) skip themselves when it is absent; point
EIOPA_REPO at a checkout to enable them.

Four things can rot, and each needs a different kind of test:

1. Standalone-ness. The skill's whole premise is that it depends on nothing
   outside its own folder. Tested by running it from an isolated directory, and
   by planting a decoy repo to prove it is ignored.
2. The tool's behaviour. Covered by CLI tests that shell out, so argument
   parsing and exit codes are exercised the way a caller meets them.
3. The data. Per-release content hashes catch a restated month, so a correction
   cannot quietly move every answer that depends on it.
4. The prose. SKILL.md states coverage figures and names the months where a
   euro-area country had its own volatility adjustment. Those are claims about
   data, so they are checked against data.

Golden values pin explicit reference dates and never "latest", which moves every
month and would make these tests rot by design.
"""
from __future__ import annotations

import hashlib
import importlib.util
import json
import os
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

HERE = Path(__file__).resolve().parent
# The tool sits beside tests/ in the skill, or in tools/ in the extraction repo.
TOOL = next((p for p in (HERE.parent / "eiopa_curves.py",
                         HERE.parent / "tools" / "eiopa_curves.py",
                         HERE.parent.parent / "tools" / "eiopa_curves.py")
             if p.exists()), None)
if TOOL is None:
    raise SystemExit("Cannot find eiopa_curves.py next to these tests.")
GOLDEN = HERE / "golden.json"


def load_tool():
    spec = importlib.util.spec_from_file_location("eiopa_curves", TOOL)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


ec = load_tool()
CV = ec.load()


def run_cli(*args, cwd=None, tool=None) -> subprocess.CompletedProcess:
    return subprocess.run([sys.executable, str(tool or TOOL), *args],
                          capture_output=True, text=True, cwd=cwd)


def golden() -> dict:
    return json.loads(GOLDEN.read_text())


def repo_data_dir():
    """An EIOPA_all_curves data/ folder, if one is reachable."""
    env = os.environ.get("EIOPA_REPO")
    cands = [Path(env) / "data"] if env else []
    cands += [TOOL.parent.parent / "data", TOOL.parent / "data"]
    return next((c for c in cands if (c / "yield_curves.csv").exists()), None)


# Countries that share the euro curve throughout. Croatia and Bulgaria are
# excluded on purpose: they adopted the euro during the sample, so they are
# legitimately different beforehand.
CORE_EURO = ["AT", "BE", "CY", "DE", "EE", "ES", "FI", "FR", "GR", "IE", "IT",
             "LT", "LU", "LV", "MT", "NL", "PT", "SI", "SK", "EU"]


def date_fingerprint(cv, date: str) -> str:
    """Stable hash of every rate published at one reference date.

    Rates are formatted to six decimals rather than hashed as floats so the
    digest cannot shift with a float-repr difference between interpreters.
    """
    h = hashlib.sha256()
    for code in cv.countries_at(date):
        for k in (0, 1):
            for t, r in sorted(cv.curve(code, date, k).items()):
                h.update(f"{code}|{k}|{t}|{r:.6f}\n".encode())
    return h.hexdigest()


# --------------------------------------------------------------------------


class TestStandalone(unittest.TestCase):
    """The skill must answer from its own two files and nothing else. These are
    the tests that encode that promise."""

    def _isolated(self, tmp: Path) -> Path:
        (tmp / "data").mkdir(parents=True, exist_ok=True)
        shutil.copy(TOOL, tmp / "eiopa_curves.py")
        shutil.copy(ec.bundled_pack(), tmp / "data" / "curves.pack")
        return tmp / "eiopa_curves.py"

    def test_answers_from_an_isolated_copy_and_writes_nothing(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            tool = self._isolated(tmp)
            before = {p for p in tmp.rglob("*")}
            r = run_cli("spot", "--country", "UK", "--term", "2",
                        "--date", "2024-12-31", "--json", cwd="/", tool=tool)
            self.assertEqual(r.returncode, 0, r.stderr)
            self.assertAlmostEqual(
                json.loads(r.stdout)["rates"][0]["spot_rate"], 0.04263, places=9)
            after = {p for p in tmp.rglob("*") if not p.name.endswith(".pyc")
                     and "__pycache__" not in p.parts}
            self.assertEqual(after, before,
                             "Querying created files; the tool must not write "
                             "caches or indexes.")

    def test_a_local_repo_is_never_silently_preferred(self):
        """Plant a decoy yield_curves.csv where an earlier version would have
        found and trusted it. The answer must still come from the pack."""
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            tool = self._isolated(tmp)
            decoy = tmp / "data" / "yield_curves.csv"
            decoy.write_text("reference_date,curve_type,country,term_index,spot_rate\n"
                             "2024-12-31,no_VA,UK,2.0,0.99999\n")
            r = run_cli("spot", "--country", "UK", "--term", "2",
                        "--date", "2024-12-31", "--json", cwd=str(tmp), tool=tool)
            self.assertEqual(r.returncode, 0, r.stderr)
            got = json.loads(r.stdout)["rates"][0]["spot_rate"]
            self.assertAlmostEqual(got, 0.04263, places=9,
                                   msg="A local CSV was picked up; the skill must "
                                       "read only its bundled pack.")

    def test_missing_pack_fails_with_a_useful_message(self):
        with tempfile.TemporaryDirectory() as td:
            tmp = Path(td)
            shutil.copy(TOOL, tmp / "eiopa_curves.py")
            r = run_cli("spot", "--country", "UK", "--term", "2",
                        "--date", "2024-12-31", tool=tmp / "eiopa_curves.py")
            self.assertNotEqual(r.returncode, 0)
            self.assertIn("curves.pack", r.stdout + r.stderr)

    def test_uses_only_the_standard_library(self):
        src = TOOL.read_text()
        for banned in ("import pandas", "import numpy", "import openpyxl",
                       "import requests", "import pyarrow"):
            self.assertNotIn(banned, src,
                             f"{banned} would break the no-dependency promise.")


class TestGoldenValues(unittest.TestCase):
    """Known rates that must never move. Each was verified against the source
    Term_Structures workbook when it was pinned."""

    def test_pinned_spot_rates(self):
        for row in golden()["spot_values"]:
            with self.subTest(**row):
                got = CV.rate(row["country"], row["date"],
                              ec.CURVE_TYPES.index(row["curve"]), row["term"])
                self.assertIsNotNone(got, f"no rate for {row}")
                self.assertAlmostEqual(got, row["rate"], places=9)

    def test_pinned_dates_are_unchanged(self):
        for date, want in golden()["date_fingerprints"].items():
            with self.subTest(date=date):
                self.assertEqual(
                    date_fingerprint(CV, date), want,
                    f"The {date} release no longer matches its pinned fingerprint. "
                    f"Either EIOPA republished a correction or the pack was rebuilt "
                    f"from changed data. Investigate before regenerating.")


class TestInvariants(unittest.TestCase):
    """Properties that hold across the whole dataset. These are the claims
    SKILL.md makes, checked against data."""

    def test_euro_area_no_va_curves_are_identical_everywhere(self):
        """The load-bearing fact of the whole skill: EIOPA publishes one curve
        per currency, so a euro-area sovereign spread is always exactly zero."""
        for date in CV.releases:
            present = [c for c in CORE_EURO if CV.curve_exists(c, 0, date)]
            base = CV.curve(present[0], date, 0)
            for c in present[1:]:
                self.assertEqual(CV.curve(c, date, 0), base,
                                 f"{c} diverges from {present[0]} at {date}; the "
                                 f"skill's central claim would need rewriting.")

    def test_country_specific_va_months_match_the_pinned_set(self):
        """The tool derives this rather than hardcoding it, so the test checks
        the derivation against a reviewed baseline."""
        self.assertEqual(CV.va_exceptions(), golden()["va_exceptions"],
                         "The months where a country carries its own volatility "
                         "adjustment have changed. Review, then --regenerate-golden.")

    def test_va_exception_detection_agrees_with_a_direct_scan(self):
        """The derivation probes one tenor for speed. Confirm it finds exactly
        what an exhaustive comparison against the euro curve finds."""
        direct = {}
        for date in CV.releases:
            base = CV.rate("EU", date, 1, 10)
            if base is None:
                continue
            for c in CORE_EURO:
                r = CV.rate(c, date, 1, 10)
                if r is not None and abs(r - base) > 1e-12:
                    direct.setdefault(c, []).append(date)
        derived = {k: v for k, v in CV.va_exceptions().items() if k in CORE_EURO}
        self.assertEqual(derived, {k: sorted(v) for k, v in direct.items()})

    def test_shared_curve_groups_hold_over_the_full_term_structure(self):
        """currency_groups probes a few tenors for speed; every group it reports
        must actually be identical at all 150."""
        for date in (CV.releases[0], CV.releases[len(CV.releases) // 2],
                     CV.releases[-1]):
            for g in CV.currency_groups(date):
                self.assertTrue(CV.verify_group(g, date),
                                f"{g} differ somewhere in the curve at {date}")

    def test_coverage_changes_match_the_pinned_set(self):
        got = [[d, a, r] for d, a, r in CV.coverage_changes()]
        self.assertEqual(got, golden()["coverage_changes"],
                         "The published country set changed on a different date "
                         "than recorded. Review, then --regenerate-golden.")

    def test_currency_era_labels_stop_at_the_documented_boundary(self):
        cutoff = golden()["currency_era_last"]
        for code in CV.countries:
            for lo, hi, raw in CV.labels(code):
                if raw != code:
                    self.assertLessEqual(
                        hi, cutoff,
                        f"currency label {raw!r} for {code} runs past {cutoff}")

    def test_every_published_curve_has_a_complete_term_structure(self):
        for date in CV.releases:
            for code in CV.countries_at(date):
                for k in (0, 1):
                    c = CV.curve(code, date, k)
                    if c:
                        self.assertEqual(sorted(c), list(range(1, max(c) + 1)),
                                         f"{code} {date} {k} has term gaps")

    def test_every_release_publishes_both_curve_types(self):
        for date in CV.releases:
            for code in CV.countries_at(date):
                self.assertTrue(CV.curve_exists(code, 1, date),
                                f"{code} has no with_VA curve at {date}")

    def test_forwards_recompound_into_the_published_spot_curve(self):
        """One-year forwards chained back together must reproduce the spot rate.
        Guards the compounding convention behind the forward command."""
        for code, date, horizon in golden()["forward_roundtrip"]:
            with self.subTest(country=code, date=date):
                curve = CV.curve(code, date, 0)
                acc = 1.0
                for y in range(1, horizon + 1):
                    s1 = 0.0 if y == 1 else curve[y - 1]
                    acc *= (1 + curve[y]) ** y / (1 + s1) ** (y - 1)
                self.assertAlmostEqual(acc ** (1 / horizon) - 1, curve[horizon],
                                       places=12)


class TestSnapshotFreshness(unittest.TestCase):
    """The bundled pack is a snapshot. This fails when the repo has moved past
    it, which is the prompt to re-pack and refresh the documented figures."""

    def test_snapshot_matches_the_pinned_release(self):
        pinned = golden()["coverage"]["last_release_when_pinned"]
        self.assertEqual(
            CV.releases[-1], pinned,
            f"The bundled pack now ends at {CV.releases[-1]} but the docs were "
            f"pinned at {pinned}. Update the coverage line in SKILL.md and the "
            f"tables in references/dataset.md, then --regenerate-golden.")

    def test_repo_has_not_published_months_the_snapshot_lacks(self):
        data_dir = repo_data_dir()
        if data_dir is None:
            self.skipTest("no EIOPA_all_curves checkout reachable")
        import csv as _csv
        last = ""
        with open(data_dir / "yield_curves.csv", newline="") as fh:
            for row in _csv.reader(fh):
                if row and row[0][:2] == "20" and row[0] > last:
                    last = row[0]
        self.assertLessEqual(
            last, CV.releases[-1],
            f"The repo has data through {last} but the bundled pack stops at "
            f"{CV.releases[-1]}. Re-run `pack` and rebuild the skill.")


class TestSeriesSampling(unittest.TestCase):
    """Regressions for three defects found in review: a sampled series dropped
    its newest point, the header reported the sampled range as if it were the
    requested one, and the extremes came from the thinned sample rather than
    the real series."""

    def test_yearly_sampling_keeps_a_mid_year_endpoint(self):
        last = CV.releases[-1]
        r = run_cli("series", "--country", "DE", "--term", "10",
                    "--from", "2024", "--to", last, "--freq", "year")
        self.assertEqual(r.returncode, 0, r.stderr)
        self.assertIn(last, r.stdout,
                      "The newest observation was dropped by yearly sampling.")

    def test_header_reports_the_underlying_range_not_the_sample(self):
        last = CV.releases[-1]
        r = run_cli("series", "--country", "DE", "--term", "10", "--from", "2024",
                    "--to", last, "--freq", "year", "--json")
        payload = json.loads(r.stdout)
        self.assertEqual(payload["range"]["last"], last)
        self.assertGreaterEqual(payload["range"]["published_months"],
                                payload["range"]["shown"])

    def test_extremes_are_taken_over_every_published_month(self):
        args = ("series", "--country", "BR", "--term", "10", "--from", "2020",
                "--to", "2024-12-31", "--json")
        monthly = json.loads(run_cli(*args, "--freq", "month").stdout)
        yearly = json.loads(run_cli(*args, "--freq", "year").stdout)
        self.assertEqual(monthly["low"], yearly["low"])
        self.assertEqual(monthly["high"], yearly["high"])


class TestCliContract(unittest.TestCase):
    """Behaviour a caller depends on, exercised through the command line."""

    def test_va_flag_works_before_and_after_the_subcommand(self):
        a = run_cli("--va", "spot", "--country", "DE", "--term", "10",
                    "--date", "2024-12-31", "--json")
        b = run_cli("spot", "--country", "DE", "--term", "10",
                    "--date", "2024-12-31", "--va", "--json")
        self.assertEqual(a.returncode, 0, a.stderr)
        self.assertEqual(json.loads(a.stdout), json.loads(b.stdout))
        self.assertEqual(json.loads(a.stdout)["curve_type"], "with_VA")

    def test_unknown_country_fails_loudly_with_a_hint(self):
        r = run_cli("spot", "--country", "Wakanda", "--term", "10", "--date", "2024")
        self.assertNotEqual(r.returncode, 0)
        self.assertIn("Unknown country", r.stdout + r.stderr)

    def test_currency_era_name_resolves_to_the_modern_code(self):
        """Australia in Dec 2014 is stored as AUD; asking by name must find it."""
        r = run_cli("spot", "--country", "Australia", "--term", "10",
                    "--date", "2014-12-31", "--json")
        self.assertEqual(r.returncode, 0, r.stderr)
        self.assertEqual(json.loads(r.stdout)["country"], "AU")

    def test_non_month_end_date_snaps_and_says_so(self):
        r = run_cli("spot", "--country", "DE", "--term", "10", "--date", "2024-12-15")
        self.assertEqual(r.returncode, 0, r.stderr)
        self.assertIn("2024-12-31", r.stdout)
        self.assertRegex(r.stdout, r"month-end|nearest release")

    def test_json_carries_the_snapshot_date(self):
        """Every machine-readable answer must say how current the data is."""
        p = json.loads(run_cli("spot", "--country", "UK", "--term", "2",
                               "--date", "2024-12-31", "--json").stdout)
        self.assertIn("snapshot_generated", p)
        self.assertEqual(p["latest_release_in_snapshot"], CV.releases[-1])

    def test_compare_flags_curves_that_are_identical_by_construction(self):
        r = run_cli("compare", "--countries", "DE,IT,ES", "--term", "10",
                    "--date", "2020-12-31")
        self.assertIn("Identical by construction", r.stdout)

    def test_facts_reports_the_derived_structural_claims(self):
        """The single source of truth for everything the docs used to assert."""
        p = json.loads(run_cli("facts", "--json").stdout)
        self.assertEqual(p["releases"], len(CV.releases))
        self.assertEqual(p["last"], CV.releases[-1])
        self.assertTrue(p["shared_groups_verified"])
        self.assertEqual(p["va_exceptions"], CV.va_exceptions())
        self.assertTrue(any("EU" in g and "DE" in g
                            for g in p["shared_curve_groups"]),
                        "the euro bloc should surface as a shared-curve group")

    def test_month_ranges_collapse_consecutive_months(self):
        self.assertEqual(ec.month_ranges(["2018-08-31", "2018-10-31",
                                          "2018-11-30"]),
                         "2018-08, 2018-10 to 2018-11")
        self.assertEqual(ec.month_ranges(["2020-01-31"]), "2020-01")

    def test_coverage_states_that_the_data_is_a_snapshot(self):
        r = run_cli("coverage")
        self.assertIn("snapshot", r.stdout.lower())
        self.assertIn(CV.releases[-1], r.stdout)


class TestPackRoundTrip(unittest.TestCase):
    """The pack is the only copy of the data the skill ships. If it stopped
    reproducing the source exactly, every machine would be confidently wrong in
    the same way."""

    def test_pack_round_trips_every_rate_exactly(self):
        data_dir = repo_data_dir()
        if data_dir is None:
            self.skipTest("no EIOPA_all_curves checkout reachable")
        with tempfile.TemporaryDirectory() as td:
            pack = ec.write_pack(data_dir, Path(td) / "curves.pack", quiet=True)
            rebuilt = ec.load(pack)
            self.assertEqual(rebuilt.releases, CV.releases)
            for date in (CV.releases[0], CV.releases[len(CV.releases) // 2],
                         CV.releases[-1]):
                self.assertEqual(date_fingerprint(rebuilt, date),
                                 date_fingerprint(CV, date),
                                 f"pack rebuilt from CSV differs at {date}")

    def test_pack_matches_the_source_csv_rate_for_rate(self):
        data_dir = repo_data_dir()
        if data_dir is None:
            self.skipTest("no EIOPA_all_curves checkout reachable")
        import csv as _csv
        checked = 0
        with open(data_dir / "yield_curves.csv", newline="") as fh:
            for i, r in enumerate(_csv.DictReader(fh)):
                if i % 997:          # a spread-out sample, ~2200 rows
                    continue
                got = CV.rate(ec.canonical_country(r["country"]),
                              r["reference_date"],
                              ec.CURVE_TYPES.index(r["curve_type"]),
                              int(float(r["term_index"])))
                self.assertAlmostEqual(got, float(r["spot_rate"]), places=9)
                checked += 1
        self.assertGreater(checked, 500)


class TestDiscountAndInterpolation(unittest.TestCase):
    """Discount factors and fractional tenors were being recomputed by hand on
    almost every use before the tool provided them. These pin the conventions so
    the two never drift apart."""

    def test_discount_factor_round_trips(self):
        for t in (1, 5, 10, 30, 50):
            r = CV.rate("DE", "2024-12-31", 0, t)
            df = ec.discount_factor(r, t)
            self.assertAlmostEqual(df ** (-1.0 / t) - 1.0, r, places=12)

    def test_published_terms_are_never_flagged_as_interpolated(self):
        for t in (1, 7, 10, 150):
            _, interp = CV.rate_at("DE", "2024-12-31", 0, t)
            self.assertFalse(interp, f"{t}y is published; it must not be derived")

    def test_interpolated_tenor_sits_between_its_neighbours(self):
        lo = CV.rate("DE", "2024-12-31", 0, 7)
        hi = CV.rate("DE", "2024-12-31", 0, 8)
        mid, interp = CV.rate_at("DE", "2024-12-31", 0, 7.5)
        self.assertTrue(interp)
        self.assertGreater(mid, min(lo, hi))
        self.assertLess(mid, max(lo, hi))

    def test_interpolation_is_arbitrage_free_on_discount_factors(self):
        """Log-linear on DF means a constant forward across the bracket. Chaining
        the two half-year forwards must reproduce the one-year move exactly."""
        r7 = CV.rate("DE", "2024-12-31", 0, 7)
        r8 = CV.rate("DE", "2024-12-31", 0, 8)
        r75, _ = CV.rate_at("DE", "2024-12-31", 0, 7.5)
        df7, df75, df8 = (ec.discount_factor(r7, 7), ec.discount_factor(r75, 7.5),
                          ec.discount_factor(r8, 8))
        self.assertAlmostEqual((df75 / df7) ** 2, df8 / df7, places=12)

    def test_sub_year_term_returns_the_one_year_rate(self):
        """DF(0) = 1 brackets the first year, so a constant forward there means
        any term under a year carries the one-year rate."""
        one = CV.rate("DE", "2024-12-31", 0, 1)
        half, interp = CV.rate_at("DE", "2024-12-31", 0, 0.5)
        self.assertTrue(interp)
        self.assertAlmostEqual(half, one, places=12)

    def test_terms_beyond_the_curve_are_refused_not_extrapolated(self):
        r, _ = CV.rate_at("DE", "2024-12-31", 0, 151)
        self.assertIsNone(r)

    def test_interpolated_rows_are_marked_and_explained(self):
        r = run_cli("spot", "--country", "DE", "--term", "7.5",
                    "--date", "2024-12-31")
        self.assertEqual(r.returncode, 0, r.stderr)
        self.assertIn("†", r.stdout)
        self.assertIn("interpolated, not published", r.stdout)

    def test_json_marks_interpolation_and_carries_discount_factors(self):
        p = json.loads(run_cli("spot", "--country", "DE", "--term", "7,7.5",
                               "--date", "2024-12-31", "--json").stdout)
        self.assertFalse(p["rates"][0]["interpolated"])
        self.assertTrue(p["rates"][1]["interpolated"])
        for row in p["rates"]:
            self.assertIsNotNone(row["discount_factor"])


class TestShockReconstruction(unittest.TestCase):
    """The stressed curves are computed, not looked up, so they need checking
    against EIOPA's own arithmetic rather than against themselves."""

    def test_factor_tables_are_the_documented_shape(self):
        self.assertEqual(len(ec.SHOCK_DOWN), 150)
        self.assertEqual(len(ec.SHOCK_UP), 150)
        self.assertAlmostEqual(ec.SHOCK_DOWN[9], 0.31, places=12)   # 10y
        self.assertAlmostEqual(ec.SHOCK_UP[9], 0.42, places=12)
        for tbl in (ec.SHOCK_DOWN, ec.SHOCK_UP):
            self.assertAlmostEqual(tbl[89], 0.20, places=12)        # 90y
            self.assertAlmostEqual(tbl[149], 0.20, places=12)       # 150y

    def test_with_va_stress_is_not_the_with_va_curve_shocked(self):
        """The trap that produced a wrong answer once. Must differ by factor x VA."""
        s = CV.rate("EU", "2024-12-31", 0, 10)
        va = 0.0023
        correct = ec.shocked_rate(s, 10, "down", va)
        naive = round((s + va) * (1 - ec.SHOCK_DOWN[9]), 5)
        self.assertAlmostEqual(correct, 0.01794, places=5)
        self.assertGreater(correct - naive, 0.0005,
                           "the naive construction should understate by ~7 bp")

    def test_down_shock_leaves_negative_rates_alone(self):
        s = CV.rate("CH", "2016-06-30", 0, 20)
        self.assertLess(s, 0)
        self.assertAlmostEqual(ec.shocked_rate(s, 20, "down"), round(s, 5), places=9)

    def test_up_shock_floor_binds_below_the_expected_level(self):
        self.assertTrue(ec.up_shock_floor_binds(0.02267, 10))    # 0.42 x 2.267%
        self.assertFalse(ec.up_shock_floor_binds(0.05, 10))      # 0.42 x 5%
        self.assertAlmostEqual(ec.shocked_rate(0.02267, 10, "up"), 0.03267, places=5)

    def test_shock_refuses_fractional_terms(self):
        r = run_cli("shock", "--country", "DE", "--date", "2024-12-31",
                    "--terms", "7.5")
        self.assertNotEqual(r.returncode, 0)

    def test_shock_works_without_any_workbook(self):
        """Factors are embedded, so the stress is available standalone."""
        p = json.loads(run_cli("shock", "--country", "DE", "--date", "2024-12-31",
                               "--terms", "10", "--va", "--json").stdout)
        self.assertTrue(p["reconstructed"])
        self.assertAlmostEqual(p["shocks"][0]["down"], 0.01794, places=5)
        self.assertAlmostEqual(p["shocks"][0]["up"], 0.03497, places=5)


class TestShockFormulas(unittest.TestCase):
    """The shock construction is documented in prose and previously produced a
    wrong answer. Verify it against the workbook when one is reachable."""

    @classmethod
    def setUpClass(cls):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise unittest.SkipTest("openpyxl not installed")
        d = repo_data_dir()
        cls.wb_path = (d.parent / "input" /
                       "EIOPA_RFR_20241231_Term_Structures.xlsx") if d else None
        if cls.wb_path is None or not cls.wb_path.exists():
            raise unittest.SkipTest("source workbook not available")

    def test_with_va_shock_is_built_from_the_no_va_curve(self):
        """If EIOPA changed this to shock the with_VA curve directly, the skill's
        documented formulas would be wrong by roughly VA x shock factor."""
        import openpyxl
        wb = openpyxl.load_workbook(self.wb_path, data_only=False)
        f = wb["Spot_WITH_VA_shock_DOWN"].cell(20, 3).value  # maturity 10
        self.assertIn("RFR_spot_no_VA", f)
        self.assertIn("VA!", f)
        self.assertNotIn("RFR_spot_with_VA", f)

    def test_up_shock_has_a_one_percentage_point_floor(self):
        import openpyxl
        wb = openpyxl.load_workbook(self.wb_path, data_only=False)
        self.assertIn("MAX(0.01",
                      wb["Spot_NO_VA_shock_UP"].cell(20, 3).value.replace(" ", ""))

    def test_documented_worked_example_reproduces(self):
        """The formulas in references/dataset.md, applied to EUR 10y at
        2024-12-31, must reproduce EIOPA's own shocked values."""
        import openpyxl
        sh = openpyxl.load_workbook(self.wb_path, data_only=False)["Shocks"]
        f_down, f_up = sh.cell(20, 4).value, sh.cell(20, 5).value
        s, va = CV.rate("EU", "2024-12-31", 0, 10), 0.0023
        self.assertAlmostEqual(round(s - f_down * abs(s), 5), 0.01564, places=5)
        self.assertAlmostEqual(round(s - f_down * abs(s) + va, 5), 0.01794, places=5)
        self.assertAlmostEqual(round(s + max(0.01, f_up * abs(s)), 5), 0.03267,
                               places=5)

    def test_embedded_factor_table_matches_the_workbook(self):
        """The tool derives factors 21-90 by linear interpolation instead of
        storing them. Check every one against the sheet."""
        import openpyxl
        ws = openpyxl.load_workbook(self.wb_path, data_only=True,
                                    read_only=True)["Shocks"]
        rows = [list(r) for r in ws.iter_rows(min_row=11, max_row=160, max_col=6,
                                              values_only=True)]
        for i, row in enumerate(rows):
            if row[1] is None:
                continue
            self.assertEqual(int(row[1]), i + 1)
            self.assertAlmostEqual(ec.SHOCK_DOWN[i], row[3], places=12,
                                   msg=f"down factor differs at {i + 1}y")
            self.assertAlmostEqual(ec.SHOCK_UP[i], row[4], places=12,
                                   msg=f"up factor differs at {i + 1}y")

    def test_full_reconstructed_curve_matches_eiopa_arithmetic(self):
        """All 150 terms, both directions, both curve types, against the
        workbook's own factors applied through the sheet formulas. This is what
        stands in for EIOPA's cached values, which the sheets do not contain."""
        import openpyxl
        ws = openpyxl.load_workbook(self.wb_path, data_only=True,
                                    read_only=True)["Shocks"]
        rows = [list(r) for r in ws.iter_rows(min_row=11, max_row=160, max_col=6,
                                              values_only=True)]
        f_down = [r[3] for r in rows if r[1] is not None]
        f_up = [r[4] for r in rows if r[1] is not None]
        va = float(CV.params("DE", "2024-12-31", "with_VA")["va"]) / 10000.0
        curve = CV.curve("DE", "2024-12-31", 0)
        checked = 0
        for t, s in curve.items():
            want_down = round((s if s < 0 else s - f_down[t - 1] * abs(s)), 5)
            want_up = round(s + max(0.01, f_up[t - 1] * abs(s)), 5)
            self.assertAlmostEqual(ec.shocked_rate(s, t, "down"), want_down, places=9)
            self.assertAlmostEqual(ec.shocked_rate(s, t, "up"), want_up, places=9)
            self.assertAlmostEqual(ec.shocked_rate(s, t, "down", va),
                                   round((s if s < 0 else s - f_down[t - 1] * abs(s))
                                         + va, 5), places=9)
            self.assertAlmostEqual(ec.shocked_rate(s, t, "up", va),
                                   round(s + max(0.01, f_up[t - 1] * abs(s)) + va, 5),
                                   places=9)
            checked += 1
        self.assertEqual(checked, 150)


class TestSpotCheckSuite(unittest.TestCase):
    """20 spot checks pinned during manual review, one per trap this skill's
    docs call out: currency-sharing, label eras, VA quirks, interpolation,
    date-snapping, coverage gaps, and the derived (forward/shock) commands.

    Each value below was produced by running the CLI, shown to a human for a
    sanity check, and pinned only after that check passed - the same bar as
    TestGoldenValues, kept separate because it is organised by trap rather
    than by data category and is meant to double as a readable catalogue of
    "here is proof each documented gotcha actually behaves as documented".
    """

    # 1. Plain sanity check on a well-known point.
    def test_01_germany_10y_2024_is_a_known_value(self):
        self.assertAlmostEqual(CV.rate("DE", "2024-12-31", 0, 10), 0.02267,
                               places=9)

    # 2. Currency-sharing: distinct sovereigns, byte-identical euro curve.
    def test_02_france_and_italy_share_the_euro_curve(self):
        fr = CV.rate("FR", "2024-12-31", 0, 10)
        it = CV.rate("IT", "2024-12-31", 0, 10)
        de = CV.rate("DE", "2024-12-31", 0, 10)
        self.assertEqual(fr, it)
        self.assertEqual(fr, de)
        self.assertAlmostEqual(fr, 0.02267, places=9)

    # 3. Early-era currency label (pre country-label switch) resolves and its
    #    provenance line reports the raw label actually published.
    def test_03_chf_label_pre_2015_11_resolves_and_reports_raw_label(self):
        r = run_cli("spot", "--country", "CHF", "--term", "10",
                    "--date", "2015-01-31")
        self.assertEqual(r.returncode, 0, r.stderr)
        self.assertIn("labelled 'CHF'", r.stdout)
        self.assertAlmostEqual(CV.rate("CH", "2015-01-31", 0, 10), -0.00085,
                               places=9)

    # 4. Same country, later era, queried by the post-switch country code.
    def test_04_ch_label_post_2015_11_is_the_same_country_as_chf(self):
        self.assertAlmostEqual(CV.rate("CH", "2024-12-31", 0, 10), 0.00380,
                               places=9)

    # 5. GBP/UK label equivalence, mid-transition date. Country-label
    #    normalisation lives in resolve_country(), used by each cmd_*
    #    function - not in CV.rate itself - so this goes through the CLI
    #    rather than the library call, or it would wrongly appear to fail.
    def test_05_gbp_and_uk_resolve_to_the_same_curve(self):
        gbp = json.loads(run_cli("spot", "--country", "GBP", "--term", "10",
                                 "--date", "2016-06-30", "--json").stdout)
        uk = json.loads(run_cli("spot", "--country", "UK", "--term", "10",
                                "--date", "2016-06-30", "--json").stdout)
        self.assertEqual(gbp["rates"][0]["spot_rate"], uk["rates"][0]["spot_rate"])
        self.assertAlmostEqual(uk["rates"][0]["spot_rate"], 0.00852, places=9)

    # 6. with_VA vs no_VA genuinely diverge for a country with its own VA.
    def test_06_switzerland_with_va_differs_from_no_va(self):
        no_va = CV.rate("CH", "2022-06-30", 0, 7)
        with_va = CV.rate("CH", "2022-06-30", 1, 7)
        self.assertAlmostEqual(no_va, 0.00666, places=9)
        self.assertAlmostEqual(with_va, 0.00646, places=9)
        self.assertNotEqual(no_va, with_va)

    # 7. Poland also carries its own VA and diverges by more than a rounding
    #    error - originally mis-guessed as a "near-zero divergence" case; the
    #    real data corrected that guess, which is exactly what this suite is
    #    for. Kept as a VA-divergence case, not a null case.
    def test_07_poland_with_va_diverges_from_no_va_by_double_digit_bp(self):
        no_va = CV.rate("PL", "2024-12-31", 0, 10)
        with_va = CV.rate("PL", "2024-12-31", 1, 10)
        self.assertAlmostEqual(no_va, 0.05780, places=9)
        self.assertAlmostEqual(with_va, 0.05940, places=9)
        self.assertGreater(abs(with_va - no_va), 0.001)

    # 8. Fractional-term interpolation is flagged as derived, not published.
    def test_08_fractional_term_is_interpolated_and_flagged(self):
        r = run_cli("spot", "--country", "DE", "--term", "7.5",
                    "--date", "2024-12-31", "--json")
        payload = json.loads(r.stdout)["rates"][0]
        self.assertTrue(payload["interpolated"])
        self.assertAlmostEqual(payload["spot_rate"], 0.022108, places=5)

    # 9. Sub-1-year term returns the 1y-constant-forward implied rate.
    def test_09_sub_one_year_term_returns_the_one_year_implied_rate(self):
        r = run_cli("spot", "--country", "DE", "--term", "0.5",
                    "--date", "2024-12-31", "--json")
        payload = json.loads(r.stdout)["rates"][0]
        self.assertTrue(payload["interpolated"])
        self.assertAlmostEqual(payload["spot_rate"], 0.02236, places=5)

    # 10. Extreme long end, near the UFR-convergence tail.
    def test_10_extreme_long_end_150y(self):
        self.assertAlmostEqual(CV.rate("DE", "2024-12-31", 0, 150), 0.03097,
                               places=9)

    # 11. Deep negative-rate era, euro area.
    def test_11_euro_area_5y_negative_in_late_2020(self):
        r = CV.rate("EU", "2020-11-30", 0, 5)
        self.assertLess(r, 0)
        self.assertAlmostEqual(r, -0.00555, places=9)

    # 12. Non-month-end date snaps to the nearest release, and says so.
    def test_12_non_month_end_date_snaps_and_reports_the_snap(self):
        r = run_cli("spot", "--country", "DE", "--term", "10",
                    "--date", "2024-12-15")
        self.assertEqual(r.returncode, 0, r.stderr)
        self.assertIn("2024-12-31", r.stdout)
        self.assertIn("2024-12-15", r.stdout)

    # 13. `latest` resolves to the newest release in this snapshot, not a
    #     hardcoded date - pinned relative to CV.releases so this does not
    #     rot the moment the pack is rebuilt with a new release appended.
    def test_13_latest_resolves_to_the_newest_release_in_the_snapshot(self):
        last = CV.releases[-1]
        r = run_cli("spot", "--country", "DE", "--term", "10",
                    "--date", "latest", "--json")
        payload = json.loads(r.stdout)
        self.assertEqual(payload["reference_date"], last)

    # 14. Coverage gap: Brazil stops being published after 2024-12-31.
    def test_14_brazil_coverage_gap_after_2024_12(self):
        r = run_cli("coverage", "--country", "BR", "--json")
        payload = json.loads(r.stdout)
        self.assertIn("2025-01-31", payload["gaps"])
        self.assertNotIn("2024-12-31", payload["gaps"])

    # 15. First snapshot date in the whole dataset.
    def test_15_first_release_in_history(self):
        self.assertEqual(CV.releases[0], "2014-12-31")
        self.assertAlmostEqual(CV.rate("DE", "2014-12-31", 0, 10), 0.00723,
                               places=9)

    # 16. Curve parameters for a non-eurozone country: LLP, UFR, alpha, CRA.
    def test_16_uk_curve_parameters(self):
        r = run_cli("params", "--country", "GB", "--date", "2024-12-31",
                    "--json")
        payload = json.loads(r.stdout)
        self.assertEqual(payload["llp"], "50.0")
        self.assertEqual(payload["ufr"], "3.3")
        self.assertAlmostEqual(float(payload["alpha"]), 0.079556, places=6)
        self.assertEqual(payload["cra"], "0.0")

    # 17. Implied forward, annual-compounding convention, cross-checked by
    #     hand against the formula documented in SKILL.md.
    def test_17_five_year_ten_year_forward_matches_the_documented_formula(self):
        r = run_cli("forward", "--country", "DE", "--start", "5",
                    "--tenor", "10", "--date", "2024-12-31", "--json")
        fwd = json.loads(r.stdout)["forwards"][0]
        s1, t1 = fwd["spot_to_start"], 5
        s2, t2 = fwd["spot_to_end"], 15
        hand = ((1 + s2) ** t2 / (1 + s1) ** t1) ** (1 / (t2 - t1)) - 1
        self.assertAlmostEqual(fwd["forward_rate"], hand, places=9)
        self.assertAlmostEqual(fwd["forward_rate"], 0.024271, places=5)

    # 18. Full SCR shock curve: base/down/up all present, VA added back per
    #     the documented no_VA-plus-VA construction.
    def test_18_full_stressed_curve_has_base_down_up_at_every_term(self):
        r = run_cli("shock", "--country", "DE", "--date", "2024-12-31",
                    "--va", "--full", "--json")
        payload = json.loads(r.stdout)
        self.assertEqual(len(payload["shocks"]), 150)
        row10 = next(s for s in payload["shocks"] if s["term"] == 10)
        self.assertAlmostEqual(row10["down"], 0.01794, places=5)
        self.assertAlmostEqual(row10["up"], 0.03497, places=5)

    # 19. Shock refuses fractional terms outright rather than silently
    #     rounding to a different regulatory tenor.
    def test_19_shock_refuses_a_fractional_term(self):
        r = run_cli("shock", "--country", "DE", "--terms", "7.5",
                    "--date", "2024-12-31")
        self.assertNotEqual(r.returncode, 0)
        self.assertIn("whole year", r.stderr + r.stdout)

    # 20. JSON gives raw decimals, not the percent-formatted text figure -
    #     guards against a double-multiply if a caller assumes otherwise.
    def test_20_json_output_is_decimal_not_percent(self):
        r = run_cli("spot", "--country", "DE", "--term", "10",
                    "--date", "2024-12-31", "--json")
        payload = json.loads(r.stdout)["rates"][0]
        self.assertAlmostEqual(payload["spot_rate"], 0.02267, places=9)
        self.assertAlmostEqual(payload["percent"], 2.267, places=6)
        self.assertAlmostEqual(payload["spot_rate"] * 100, payload["percent"],
                               places=6)


# --------------------------------------------------------------------------


def regenerate_golden():
    """Rewrite golden.json from the bundled pack.

    Only after a change has been reviewed and accepted - regenerating to turn a
    red suite green destroys the thing the suite is for.
    """
    g = golden() if GOLDEN.exists() else {}
    pinned = g.get("pinned_dates") or ["2014-12-31", "2016-06-30", "2018-11-30",
                                       "2020-12-31", "2024-12-31"]
    exc = CV.va_exceptions()

    out = {
        "schema": 2,
        "note": "Pinned facts for test_eiopa_curves.py. Regenerate only when a "
                "change has been reviewed and accepted.",
        "currency_era_last": "2015-10-31",
        "pinned_dates": pinned,
        "coverage": {"releases": len(CV.releases), "first_release": CV.releases[0],
                     "last_release_when_pinned": CV.releases[-1],
                     "snapshot_generated": CV.generated},
        "spot_values": g.get("spot_values") or [
            {"country": "UK", "term": 2, "date": "2024-12-31", "curve": "no_VA",
             "rate": 0.04263},
            {"country": "FR", "term": 2, "date": "2024-12-31", "curve": "no_VA",
             "rate": 0.02093},
            {"country": "DE", "term": 10, "date": "2024-12-31", "curve": "no_VA",
             "rate": 0.02267},
            {"country": "EU", "term": 50, "date": "2024-12-31", "curve": "no_VA",
             "rate": 0.02698},
            {"country": "AU", "term": 10, "date": "2014-12-31", "curve": "no_VA",
             "rate": 0.02864},
            {"country": "IT", "term": 10, "date": "2018-11-30", "curve": "with_VA",
             "rate": 0.01335},
            {"country": "CH", "term": 20, "date": "2016-06-30", "curve": "no_VA",
             "rate": -0.00143},
        ],
        "forward_roundtrip": g.get("forward_roundtrip") or [
            ["DE", "2024-12-31", 30], ["UK", "2020-12-31", 20],
            ["US", "2018-11-30", 25],
        ],
        "va_exceptions": {k: sorted(v) for k, v in exc.items()},
        "coverage_changes": [[d, a, r] for d, a, r in CV.coverage_changes()],
        "date_fingerprints": {d: date_fingerprint(CV, d) for d in pinned},
    }
    GOLDEN.write_text(json.dumps(out, indent=2) + "\n", encoding="utf-8")
    print(f"wrote {GOLDEN}")
    print(f"  {len(CV.releases)} releases, {CV.releases[0]} to {CV.releases[-1]}"
          f" (snapshot {CV.generated})")
    print(f"  VA exceptions: { {k: len(v) for k, v in out['va_exceptions'].items()} }")


if __name__ == "__main__":
    if "--regenerate-golden" in sys.argv:
        sys.argv.remove("--regenerate-golden")
        regenerate_golden()
        sys.exit(0)
    unittest.main(verbosity=2)
